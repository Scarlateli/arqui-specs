"""
core/spreadsheet.py — Manipulação de planilhas Excel para o Arqui Specs.

Responsabilidades:
- Criar planilha template com estrutura padrão (SCRUM-61)
- Ler/escrever workbooks em memória via BytesIO (SCRUM-50)
- Aplicar operações do Claude (adicionar/atualizar itens) (SCRUM-51)
- Formatar profissionalmente: cabeçalho, ambientes em caixa alta, cores de status (SCRUM-52)
- Exportar para preview (DataFrame) (SCRUM-54)

Estrutura da planilha:
  Linha 1: Cabeçalho do projeto (nome, cliente, responsável, emissão)
  Linha 2: Cabeçalhos das colunas
  Linha 3+: Dados (com linhas de seção por ambiente)

Colunas:
  A: AMBIENTE  B: ITEM  C: DESCRIÇÃO  D: MARCA  E: MODELO / CÓDIGO
  F: ACABAMENTO  G: BS. TÉCNICAS  H: STATUS
"""

from io import BytesIO
import json
import datetime

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
import pandas as pd


# ─── Constantes de layout ───────────────────────────────────────────────────

COLS = {
    "ambiente":   "A",
    "item":       "B",
    "descricao":  "C",
    "marca":      "D",
    "modelo":     "E",
    "acabamento": "F",
    "bs_tecnicas":"G",
    "status":     "H",
    "_id":        "I",   # Coluna oculta: ID único para atualização via Claude
}

COL_WIDTHS = {
    "A": 18,   # AMBIENTE
    "B": 20,   # ITEM
    "C": 30,   # DESCRIÇÃO
    "D": 16,   # MARCA
    "E": 22,   # MODELO
    "F": 18,   # ACABAMENTO
    "G": 40,   # BS. TÉCNICAS (mais largo — conteúdo técnico)
    "H": 12,   # STATUS
    "I": 5,    # ID (oculta)
}

# Linha onde começam os dados (após cabeçalho do projeto + cabeçalho de colunas)
HEADER_ROWS = 2
DATA_START_ROW = HEADER_ROWS + 1


# ─── Paleta de cores (espelha a paleta do config.toml) ──────────────────────

class Colors:
    # Cabeçalho do projeto
    PROJECT_HEADER_BG   = "1C1C1C"   # Quase preto
    PROJECT_HEADER_FONT = "FAFAF7"   # Branco quente

    # Cabeçalho de colunas
    COL_HEADER_BG   = "3D3530"       # Marrom escuro
    COL_HEADER_FONT = "F0EDE6"       # Bege claro

    # Linhas de seção de ambiente
    SECTION_BG   = "8B6F47"         # Bronze quente
    SECTION_FONT = "FAFAF7"         # Branco quente

    # Linhas de dados (alternadas)
    ROW_ODD  = "FAFAF7"             # Branco quente
    ROW_EVEN = "F0EDE6"             # Bege suave

    # Status
    STATUS_APROVADO = "4A7A5A"      # Verde sage
    STATUS_DECIDIR  = "C4904A"      # Âmbar
    STATUS_VERIFICAR= "7A6B4A"      # Marrom médio
    STATUS_FONT     = "FAFAF7"      # Branco para todos os status


# ─── Helpers de estilo ────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="1C1C1C", size=10) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")

def _border_thin() -> Border:
    thin = Side(style="thin", color="DDD9D2")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _align(horizontal="left", wrap=True) -> Alignment:
    return Alignment(horizontal=horizontal, vertical="center", wrap_text=wrap)


# ─── Criação do template base ─────────────────────────────────────────────

def create_template(
    project_name: str = "Memorial Descritivo",
    client_name: str = "(definir)",
    responsible: str = "Gabriela Lendecker · CAU A123456-7",
) -> BytesIO:
    """
    Cria uma planilha template com estrutura padrão.

    Retorna BytesIO pronto para salvar ou exibir.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Memorial"

    # ── Linha 1: Cabeçalho do projeto ──────────────────────────────────────
    ws.merge_cells("A1:H1")
    ws["A1"] = project_name
    ws["A1"].font = _font(bold=True, color=Colors.PROJECT_HEADER_FONT, size=13)
    ws["A1"].fill = _fill(Colors.PROJECT_HEADER_BG)
    ws["A1"].alignment = _align(horizontal="center", wrap=False)

    # Linha de meta-dados: cliente, responsável, data
    # (será preenchida na segunda aba "Info" — ou nas células I1:K1 fora da tabela)
    # Para simplificar, adicionar na linha 1 como sub-título via linha 2

    ws.row_dimensions[1].height = 32

    # ── Linha 2: Cabeçalhos das colunas ────────────────────────────────────
    headers = [
        ("A2", "AMBIENTE"),
        ("B2", "ITEM"),
        ("C2", "DESCRIÇÃO"),
        ("D2", "MARCA"),
        ("E2", "MODELO / CÓDIGO"),
        ("F2", "ACABAMENTO"),
        ("G2", "BS. TÉCNICAS"),
        ("H2", "STATUS"),
    ]
    for cell_addr, label in headers:
        cell = ws[cell_addr]
        cell.value = label
        cell.font = _font(bold=True, color=Colors.COL_HEADER_FONT, size=10)
        cell.fill = _fill(Colors.COL_HEADER_BG)
        cell.alignment = _align(horizontal="center", wrap=False)
        cell.border = _border_thin()

    ws.row_dimensions[2].height = 22

    # ── Ocultar coluna I (ID interno) ──────────────────────────────────────
    ws.column_dimensions["I"].hidden = True

    # ── Larguras das colunas ───────────────────────────────────────────────
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # ── Freeze panes: fixar as 2 primeiras linhas ──────────────────────────
    ws.freeze_panes = "A3"

    # ── Aba de informações do projeto ──────────────────────────────────────
    ws_info = wb.create_sheet("Info")
    ws_info["A1"] = "Nome do Projeto"
    ws_info["B1"] = project_name
    ws_info["A2"] = "Cliente"
    ws_info["B2"] = client_name
    ws_info["A3"] = "Responsável"
    ws_info["B3"] = responsible
    ws_info["A4"] = "Emissão"
    ws_info["B4"] = datetime.date.today().strftime("%d %b %Y")
    ws_info["A5"] = "Versão"
    ws_info["B5"] = "v1"
    ws_info.column_dimensions["A"].width = 20
    ws_info.column_dimensions["B"].width = 40

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ─── Load / Save ─────────────────────────────────────────────────────────

def load_workbook_from_bytes(data: bytes) -> openpyxl.Workbook:
    """Carrega um workbook a partir de bytes."""
    return openpyxl.load_workbook(BytesIO(data))


def save_workbook_to_bytes(wb: openpyxl.Workbook) -> bytes:
    """Salva um workbook em bytes."""
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─── Leitura de dados ────────────────────────────────────────────────────

def get_sheet_data(wb: openpyxl.Workbook) -> list[dict]:
    """
    Retorna os dados da planilha como lista de dicionários.
    Pula linhas de seção (células de ambiente em caixa alta, coluna B vazia).
    Usado para dar contexto ao Claude a cada turno.
    """
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=False):
        # Pular linhas completamente vazias
        if not any(cell.value for cell in row[:8]):
            continue
        # Pular linhas de seção (B vazio significa linha de ambiente)
        if row[0].value and not row[1].value:
            continue

        row_id = row[8].value if len(row) > 8 else None  # coluna I = ID
        rows.append({
            "id":          row_id,
            "ambiente":    row[0].value or "",
            "item":        row[1].value or "",
            "descricao":   row[2].value or "",
            "marca":       row[3].value or "",
            "modelo":      row[4].value or "",
            "acabamento":  row[5].value or "",
            "bs_tecnicas": row[6].value or "",
            "status":      row[7].value or "",
        })
    return rows


def get_compact_context(wb: openpyxl.Workbook) -> str:
    """
    Retorna resumo compacto da planilha para incluir em cada mensagem ao Claude.
    Mantém o contexto pequeno mas suficiente para o modelo entender o estado atual.
    """
    data = get_sheet_data(wb)
    if not data:
        return "📋 Planilha vazia — pronta para receber itens."

    # Agrupar por ambiente
    by_amb: dict[str, list[str]] = {}
    for item in data:
        amb = item["ambiente"] or "SEM AMBIENTE"
        if amb not in by_amb:
            by_amb[amb] = []
        status_icon = {"APROVADO": "✓", "DECIDIR": "?", "VERIFICAR": "!"}.get(item["status"], "·")
        by_amb[amb].append(f"{item['item']} [{status_icon}]")

    total = len(data)
    lines = [f"📋 Planilha atual: {total} {'item' if total == 1 else 'itens'}"]
    for amb, items in by_amb.items():
        lines.append(f"  • {amb}: {', '.join(items)}")

    return "\n".join(lines)


def get_project_info(wb: openpyxl.Workbook) -> dict:
    """Lê metadados da aba Info."""
    try:
        ws = wb["Info"]
        return {
            "nome":        ws["B1"].value or "Memorial Descritivo",
            "cliente":     ws["B2"].value or "(definir)",
            "responsavel": ws["B3"].value or "Gabriela Lendecker",
            "emissao":     ws["B4"].value or "",
            "versao":      ws["B5"].value or "v1",
        }
    except Exception:
        return {"nome": "Memorial Descritivo", "cliente": "", "responsavel": "", "emissao": "", "versao": "v1"}


def update_project_info(wb: openpyxl.Workbook, **kwargs) -> None:
    """Atualiza metadados do projeto na aba Info e no cabeçalho da planilha."""
    try:
        ws_info = wb["Info"]
    except KeyError:
        ws_info = wb.create_sheet("Info")

    mapping = {
        "nome":        ("B1", "A1"),
        "cliente":     ("B2", None),
        "responsavel": ("B3", None),
        "versao":      ("B5", None),
    }
    for field, (info_cell, _) in mapping.items():
        if field in kwargs:
            ws_info[info_cell] = kwargs[field]

    # Atualizar data de emissão
    ws_info["B4"] = datetime.date.today().strftime("%d %b %Y")

    # Atualizar nome no cabeçalho da planilha principal
    if "nome" in kwargs:
        ws = wb.active
        ws["A1"] = kwargs["nome"]


# ─── Operações (aplicadas via chamada de ferramenta do Claude) ────────────

def _next_id(wb: openpyxl.Workbook) -> int:
    """Gera um ID único sequencial para uma nova linha."""
    ws = wb.active
    max_id = 0
    for row in ws.iter_rows(min_row=DATA_START_ROW, min_col=9, max_col=9, values_only=True):
        if row[0] and isinstance(row[0], int):
            max_id = max(max_id, row[0])
    return max_id + 1


def _find_last_row_of_ambiente(ws, ambiente: str) -> int | None:
    """Retorna o número da última linha que pertence a um ambiente."""
    last = None
    for row in ws.iter_rows(min_row=DATA_START_ROW, max_col=1, values_only=True):
        if row[0] == ambiente:
            last = ws.iter_rows(
                min_row=DATA_START_ROW,
                max_col=1
            ).__next__()  # não ideal, vamos usar abordagem direta abaixo
    return last


def _find_section_and_last_data_row(ws, ambiente: str) -> tuple[int | None, int]:
    """
    Retorna (linha_da_secao_de_ambiente, ultima_linha_de_dados_do_ambiente).
    Se o ambiente não existe, retorna (None, ultima_linha_geral).
    """
    section_row = None
    last_data_row = 0

    for row_num in range(DATA_START_ROW, ws.max_row + 1):
        col_a = ws.cell(row=row_num, column=1).value
        col_b = ws.cell(row=row_num, column=2).value

        if col_a and not col_b:
            # Linha de seção
            if col_a.strip().upper() == ambiente.strip().upper():
                section_row = row_num
        elif col_a == ambiente or (section_row and col_b):
            last_data_row = row_num

    return section_row, last_data_row


def _apply_row_style(ws, row_num: int, is_section: bool = False, status: str = "") -> None:
    """Aplica estilos a uma linha (seção de ambiente ou linha de dados)."""
    if is_section:
        for col_idx in range(1, 9):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.fill = _fill(Colors.SECTION_BG)
            cell.font = _font(bold=True, color=Colors.SECTION_FONT, size=10)
            cell.alignment = _align(horizontal="left", wrap=False)
            cell.border = _border_thin()
    else:
        # Linha de dados
        row_count = 0
        for r in range(DATA_START_ROW, row_num):
            if ws.cell(row=r, column=2).value:
                row_count += 1
        bg = Colors.ROW_ODD if row_count % 2 == 0 else Colors.ROW_EVEN

        for col_idx in range(1, 9):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.fill = _fill(bg)
            cell.font = _font()
            cell.alignment = _align()
            cell.border = _border_thin()

        # Coluna A (ambiente) em cinza mais claro nas linhas de dados
        ws.cell(row=row_num, column=1).font = _font(color="7A7068")

        # Estilizar célula de STATUS
        status_cell = ws.cell(row=row_num, column=8)
        status_colors = {
            "APROVADO":  Colors.STATUS_APROVADO,
            "DECIDIR":   Colors.STATUS_DECIDIR,
            "VERIFICAR": Colors.STATUS_VERIFICAR,
        }
        if status in status_colors:
            status_cell.fill = _fill(status_colors[status])
            status_cell.font = _font(bold=True, color=Colors.STATUS_FONT, size=9)
            status_cell.alignment = _align(horizontal="center", wrap=False)


def apply_operations(wb: openpyxl.Workbook, operacoes: list[dict]) -> dict:
    """
    Aplica a lista de operações retornadas pelo Claude na planilha.

    Retorna dicionário com resultado das operações para feedback ao Claude.
    """
    ws = wb.active
    adicionados = []
    atualizados = []
    erros = []

    for op in operacoes:
        acao = op.get("acao")

        if acao == "adicionar_item":
            try:
                ambiente = (op.get("ambiente") or "").strip().upper()
                if not ambiente:
                    erros.append("adicionar_item sem ambiente especificado")
                    continue

                # Verificar se a seção do ambiente já existe
                section_row, last_row = _find_section_and_last_data_row(ws, ambiente)

                # Se a seção não existe, criar linha de seção de ambiente
                if section_row is None:
                    insert_at = (last_row + 1) if last_row >= DATA_START_ROW else ws.max_row + 1
                    if insert_at < DATA_START_ROW:
                        insert_at = DATA_START_ROW

                    # Inserir linha de seção
                    ws.insert_rows(insert_at)
                    ws.cell(row=insert_at, column=1).value = ambiente
                    ws.cell(row=insert_at, column=2).value = None
                    _apply_row_style(ws, insert_at, is_section=True)
                    section_row = insert_at
                    last_row = insert_at

                # Inserir linha de dados após a última linha do ambiente
                data_row = last_row + 1
                ws.insert_rows(data_row)

                row_id = _next_id(wb)
                status = op.get("status", "APROVADO")

                ws.cell(row=data_row, column=1).value = ambiente
                ws.cell(row=data_row, column=2).value = op.get("item", "")
                ws.cell(row=data_row, column=3).value = op.get("descricao", "")
                ws.cell(row=data_row, column=4).value = op.get("marca", "")
                ws.cell(row=data_row, column=5).value = op.get("modelo", "")
                ws.cell(row=data_row, column=6).value = op.get("acabamento", "")
                ws.cell(row=data_row, column=7).value = op.get("bs_tecnicas", "")
                ws.cell(row=data_row, column=8).value = status
                ws.cell(row=data_row, column=9).value = row_id  # ID oculto

                _apply_row_style(ws, data_row, is_section=False, status=status)
                ws.row_dimensions[data_row].height = 35

                adicionados.append(f"{ambiente} · {op.get('item', '?')}")

            except Exception as e:
                erros.append(f"adicionar_item erro: {str(e)}")

        elif acao == "atualizar_item":
            linha_id = op.get("linha_id")
            if not linha_id:
                erros.append("atualizar_item sem linha_id")
                continue

            # Encontrar a linha com esse ID
            target_row = None
            for row_num in range(DATA_START_ROW, ws.max_row + 1):
                if ws.cell(row=row_num, column=9).value == linha_id:
                    target_row = row_num
                    break

            if target_row is None:
                erros.append(f"atualizar_item: linha ID {linha_id} não encontrada")
                continue

            # Atualizar apenas os campos fornecidos
            field_map = {
                "ambiente":    1,
                "item":        2,
                "descricao":   3,
                "marca":       4,
                "modelo":      5,
                "acabamento":  6,
                "bs_tecnicas": 7,
                "status":      8,
            }
            for field, col_idx in field_map.items():
                if field in op and op[field] is not None:
                    ws.cell(row=target_row, column=col_idx).value = op[field]

            status = ws.cell(row=target_row, column=8).value or ""
            _apply_row_style(ws, target_row, is_section=False, status=status)
            atualizados.append(f"ID {linha_id} · {op.get('item', '?')}")

        else:
            erros.append(f"ação desconhecida: {acao}")

    # Atualizar data de emissão
    try:
        wb["Info"]["B4"] = datetime.date.today().strftime("%d %b %Y")
        info = get_project_info(wb)
        _update_version(wb)
    except Exception:
        pass

    return {
        "items_adicionados": len(adicionados),
        "items_atualizados": len(atualizados),
        "adicionados":       adicionados,
        "atualizados":       atualizados,
        "erros":             erros,
        "total_itens":       len(get_sheet_data(wb)),
    }


def _update_version(wb: openpyxl.Workbook) -> None:
    """Incrementa o contador de revisão na aba Info."""
    try:
        ws_info = wb["Info"]
        version_str = ws_info["B5"].value or "v1"
        # Extrair número e incrementar
        num = int(version_str.replace("v", "").replace("r", "").strip())
        ws_info["B5"] = f"v{num + 1}"
    except Exception:
        pass  # Se falhar, não é crítico


# ─── Preview para Streamlit (SCRUM-54) ────────────────────────────────────

def get_preview_dataframe(wb: openpyxl.Workbook) -> pd.DataFrame:
    """
    Converte a planilha em DataFrame para exibição no Streamlit.
    Retorna apenas colunas visíveis (sem ID interno).
    """
    data = get_sheet_data(wb)
    if not data:
        return pd.DataFrame(columns=["AMBIENTE", "ITEM", "MARCA", "MODELO", "ACABAMENTO", "BS. TÉCNICAS", "STATUS"])

    df = pd.DataFrame(data)
    df = df.rename(columns={
        "ambiente":    "AMBIENTE",
        "item":        "ITEM",
        "descricao":   "DESCRIÇÃO",
        "marca":       "MARCA",
        "modelo":      "MODELO",
        "acabamento":  "ACABAMENTO",
        "bs_tecnicas": "BS. TÉCNICAS",
        "status":      "STATUS",
    })
    # Remover coluna interna
    if "id" in df.columns:
        df = df.drop(columns=["id"])

    return df[["AMBIENTE", "ITEM", "DESCRIÇÃO", "MARCA", "MODELO", "ACABAMENTO", "BS. TÉCNICAS", "STATUS"]]


def style_preview_dataframe(df: pd.DataFrame) -> pd.io.formats.style.Styler:
    """Aplica cores de status ao DataFrame para exibição no Streamlit."""

    def color_status(val: str) -> str:
        colors = {
            "APROVADO":  "background-color: #4A7A5A; color: #FAFAF7; font-weight: bold; border-radius: 3px; padding: 2px 6px;",
            "DECIDIR":   "background-color: #C4904A; color: #FAFAF7; font-weight: bold; border-radius: 3px; padding: 2px 6px;",
            "VERIFICAR": "background-color: #7A6B4A; color: #FAFAF7; font-weight: bold; border-radius: 3px; padding: 2px 6px;",
        }
        return colors.get(str(val).strip().upper(), "")

    styler = df.style.applymap(color_status, subset=["STATUS"])
    styler = styler.set_properties(**{"font-size": "12px"})
    styler = styler.set_table_styles([
        {"selector": "th", "props": [
            ("background-color", "#3D3530"),
            ("color", "#F0EDE6"),
            ("font-weight", "bold"),
            ("font-size", "12px"),
        ]},
        {"selector": "tr:nth-child(even)", "props": [("background-color", "#F0EDE6")]},
        {"selector": "tr:nth-child(odd)",  "props": [("background-color", "#FAFAF7")]},
    ])
    return styler
