"""
app.py — Arqui Specs · Assistente de Memorial Descritivo
Estudio GL · Gabriela Lendecker · 2026

Estrutura:
  - Auth: senha simples via st.secrets (SCRUM-58)
  - Sidebar: projeto, upload/download, info do projeto
  - Layout 2 colunas: chat (esq) + preview planilha (dir) (SCRUM-54)
  - Streaming + tool use (SCRUM-49, 51, 60)
  - Auto-save draft (SCRUM-62)
  - Tratamento de erros amigável (SCRUM-55)
"""

import streamlit as st
import openpyxl
import datetime
import base64
from io import BytesIO
from pathlib import Path

from core.spreadsheet import (
    create_template,
    load_workbook_from_bytes,
    save_workbook_to_bytes,
    get_preview_dataframe,
    style_preview_dataframe,
    get_project_info,
    update_project_info,
    get_sheet_data,
)
from core.claude_client import (
    stream_response,
    build_user_message,
    MODEL,
)

# ─── Configuração da página ──────────────────────────────────────────────────

st.set_page_config(
    page_title="Arqui Specs · Estudio GL",
    page_icon="GL",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── CSS customizado ─────────────────────────────────────────────────────────

st.markdown("""
<style>
:root {
  --color-bordo:   #3D1812;
  --color-accent:  #7A2E2A;
  --color-sand:    #C8B48A;
  --color-cream:   #F7F1E8;
  --color-bg:      #FAF6EF;
  --color-ink:     #2A1510;
  --color-muted:   #6B4F45;
  --color-border:  rgba(61, 24, 18, 0.12);
  --color-sidebar-border: rgba(217, 203, 168, 0.18);
}

/* ── Área principal: clara e leve ── */
html, body,
.stApp,
[data-testid="stAppViewContainer"],
[data-testid="stMainBlockContainer"],
.main .block-container {
    background-color: var(--color-bg) !important;
    color: var(--color-ink);
}

/* ── Sidebar: bordo (identidade visual) ── */
[data-testid="stSidebar"] {
    min-width: 260px;
    max-width: 280px;
    background-color: var(--color-bordo) !important;
    border-right: 1px solid var(--color-sidebar-border);
}
[data-testid="stSidebar"] * { color: var(--color-cream); }
[data-testid="stSidebar"] hr { border-color: var(--color-sidebar-border); }

/* ── Inputs da sidebar ── */
[data-testid="stSidebar"] input,
[data-testid="stSidebar"] textarea,
[data-testid="stSidebar"] [data-baseweb="input"] > div {
    background-color: rgba(250, 246, 239, 0.12) !important;
    color: var(--color-cream) !important;
    border: 1px solid var(--color-sidebar-border) !important;
}
[data-testid="stSidebar"] input::placeholder { color: rgba(247,241,232,0.45) !important; }

/* ── Botões sidebar ── */
[data-testid="stSidebar"] .stButton > button {
    background-color: #5C2018 !important;
    color: #F7F1E8 !important;
    border: 1px solid rgba(247,241,232,0.25) !important;
    font-weight: 500 !important;
}
[data-testid="stSidebar"] .stButton > button *,
[data-testid="stSidebar"] .stButton > button p,
[data-testid="stSidebar"] .stButton > button span {
    color: #F7F1E8 !important;
    background: transparent !important;
}
[data-testid="stSidebar"] .stButton > button:hover {
    background-color: #7A2E2A !important;
    border-color: rgba(247,241,232,0.50) !important;
}
[data-testid="stSidebar"] [data-testid="stDownloadButton"] button {
    background-color: var(--color-accent) !important;
    color: var(--color-cream) !important;
    border: none !important;
    font-weight: 600 !important;
}
[data-testid="stSidebar"] [data-testid="stDownloadButton"] button:hover {
    background-color: #8F3B36 !important;
}

/* ── File uploader da sidebar ── */
[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] {
    background-color: rgba(247,241,232,0.08) !important;
    border: 1px dashed rgba(247,241,232,0.30) !important;
    border-radius: 6px !important;
}
[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] * {
    color: var(--color-cream) !important;
}
[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button {
    background-color: transparent !important;
    border: 1px solid rgba(247,241,232,0.45) !important;
    color: var(--color-cream) !important;
}

/* ── Branding header ── */
.brand-header {
    display: flex;
    align-items: center;
    gap: 12px;
    margin-bottom: 16px;
}
.brand-logo {
    height: 40px;
    width: 40px;
    display: block;
    border-radius: 6px;
    object-fit: cover;
}
.brand-name {
    font-family: Georgia, "Times New Roman", serif;
    font-size: 14px;
    font-weight: 400;
    letter-spacing: 0.06em;
    text-transform: uppercase;
    color: var(--color-cream);
    line-height: 1.2;
}
.brand-product {
    font-size: 19px;
    font-weight: 700;
    color: var(--color-sand);
    letter-spacing: 0.03em;
    margin-bottom: 1px;
}
.brand-sub {
    font-size: 10px;
    color: rgba(200,180,138,0.75);
    letter-spacing: 0.14em;
    text-transform: uppercase;
    margin-bottom: 14px;
}

/* ── Labels e texto da área principal ── */
label,
[data-testid="stWidgetLabel"] p,
[data-testid="stWidgetLabel"] span,
.stTextInput label,
.stFileUploader label,
.stSelectbox label {
    color: var(--color-ink) !important;
    opacity: 1 !important;
}

/* ── Inputs área principal ── */
input, textarea,
[data-baseweb="input"] > div,
[data-baseweb="textarea"] > div,
[data-testid="stTextInput"] input,
[data-testid="stChatInput"] textarea {
    background-color: #fff !important;
    color: var(--color-ink) !important;
    border: 1px solid var(--color-border) !important;
}
input::placeholder, textarea::placeholder {
    color: rgba(42,21,16,0.38) !important;
}

/* ── Botões área principal ── */
.stButton > button,
[data-testid="stDownloadButton"] button,
button[kind="primary"] {
    background-color: var(--color-accent) !important;
    color: var(--color-cream) !important;
    border: none !important;
}
.stButton > button:hover,
[data-testid="stDownloadButton"] button:hover {
    background-color: var(--color-bordo) !important;
}

/* ── Chat: balão assistente ── */
[data-testid="stChatMessage"]:has([data-testid="stChatMessageContent"]),
[data-testid="stChatMessage"]:has([data-testid="chatAvatarIcon-assistant"]) {
    background-color: #fff;
    border: 1px solid var(--color-border);
    border-left: 3px solid var(--color-sand);
    border-radius: 10px;
    margin-bottom: 8px;
    color: var(--color-ink);
}
[data-testid="stChatMessage"]:has([data-testid="stChatMessageContent"]) *,
[data-testid="stChatMessage"]:has([data-testid="chatAvatarIcon-assistant"]) * {
    color: var(--color-ink);
}

/* ── Chat: balão usuário ── */
[data-testid="stChatMessage"]:has([data-testid="stChatMessageAvatarUser"]),
[data-testid="stChatMessage"]:has([data-testid="chatAvatarIcon-user"]),
[data-testid="stChatMessage"]:has([aria-label*="user"]) {
    background-color: var(--color-cream) !important;
    border: 1px solid var(--color-border);
    border-left: 3px solid rgba(61,24,18,0.3);
    border-radius: 10px;
    margin-bottom: 8px;
}
[data-testid="stChatMessage"]:has([data-testid="stChatMessageAvatarUser"]) *,
[data-testid="stChatMessage"]:has([data-testid="chatAvatarIcon-user"]) *,
[data-testid="stChatMessage"]:has([aria-label*="user"]) * {
    color: var(--color-ink) !important;
}

/* ── Container chat (apenas área principal) ── */
[data-testid="stMainBlockContainer"] [data-testid="stVerticalBlockBorderWrapper"] {
    background-color: var(--color-bg) !important;
}

/* ── Sidebar: reset de containers internos ── */
[data-testid="stSidebar"] [data-testid="stVerticalBlockBorderWrapper"],
[data-testid="stSidebar"] [data-testid="column"],
[data-testid="stSidebar"] .stColumn,
[data-testid="stSidebar"] [data-testid="stHorizontalBlock"],
[data-testid="stSidebar"] div[class*="block-container"],
[data-testid="stSidebar"] [data-testid="element-container"],
[data-testid="stSidebar"] div[class*="stColumn"] {
    background-color: transparent !important;
}

/* ── Labels sidebar (override da regra global ink) ── */
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] [data-testid="stWidgetLabel"] *,
[data-testid="stSidebar"] [data-testid="stWidgetLabel"] p {
    color: rgba(247,241,232,0.75) !important;
    opacity: 1 !important;
}

/* ── Caption/rodapé da sidebar ── */
[data-testid="stSidebar"] [data-testid="stCaptionContainer"] *,
[data-testid="stSidebar"] small,
[data-testid="stSidebar"] .stCaption * {
    color: rgba(247,241,232,0.55) !important;
    opacity: 1 !important;
}

/* ── Títulos ── */
h1, h2, h3, h4 { color: var(--color-bordo) !important; }

/* ── Dataframe ── */
[data-testid="stDataFrame"] {
    border-radius: 8px;
}
/* Garantir contraste de texto nas células do dataframe */
[data-testid="stDataFrame"] [data-testid="glideDataEditor"] * {
    color: var(--color-ink) !important;
}
/* Header do dataframe: bordo com texto cream */
[data-testid="stDataFrame"] .dvn-canvas {
    background-color: var(--color-bg);
}
[data-testid="stAlert"] {
    border-left: 3px solid var(--color-accent) !important;
}

/* ── Chat container: scroll vertical fixo ── */
[data-testid="stVerticalBlockBorderWrapper"] > div {
    overflow-y: auto !important;
    overflow-x: hidden !important;
}
/* Garantir que o container de chat não cresce além do viewport */
[data-testid="stVerticalBlockBorderWrapper"] {
    max-height: inherit;
}

hr { border-color: var(--color-border); }

@media (max-width: 768px) {
    .brand-logo { height: 28px; width: 28px; }
    .brand-name { font-size: 12px; }
}

footer { visibility: hidden; }

/* ── Header e sidebar colapsado ── */
[data-testid="stHeader"],
[data-testid="stToolbar"],
[data-testid="stDecoration"] {
    background-color: transparent !important;
    border: none !important;
    box-shadow: none !important;
}
[data-testid="collapsedControl"],
[data-testid="stSidebarCollapsedControl"] {
    background-color: transparent !important;
    border: none !important;
    box-shadow: none !important;
}
[data-testid="collapsedControl"] button,
[data-testid="stSidebarCollapsedControl"] button,
[data-testid="collapsedControl"] svg,
[data-testid="stSidebarCollapsedControl"] svg {
    color: var(--color-bordo) !important;
    background-color: transparent !important;
}
</style>
""", unsafe_allow_html=True)


def _logo_data_uri() -> str:
    logo_path = Path(__file__).parent / "assets" / "logo-gl-monogram.png"
    if not logo_path.exists():
        return ""
    encoded = base64.b64encode(logo_path.read_bytes()).decode("utf-8")
    return f"data:image/png;base64,{encoded}"


# ─── Auth (SCRUM-58) ─────────────────────────────────────────────────────────

def check_auth() -> bool:
    """
    Autenticação por senha simples.
    Senha definida em .streamlit/secrets.toml como APP_PASSWORD.
    Retorna True se autenticado.
    """
    # Se não há APP_PASSWORD configurada, permite acesso (dev local)
    try:
        expected = st.secrets["APP_PASSWORD"]
    except (KeyError, FileNotFoundError):
        return True  # Sem senha configurada = acesso liberado (modo dev)

    if st.session_state.get("authenticated"):
        return True

    with st.container():
        st.markdown("## Arqui Specs")
        st.markdown("Ferramenta interna · Estudio GL")
        password = st.text_input("Senha de acesso", type="password", key="pw_input")
        if st.button("Entrar", type="primary"):
            if password == expected:
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.error("Senha incorreta.")
    return False


def get_api_key() -> str | None:
    """Retorna a API key da Anthropic dos secrets, ou None."""
    try:
        return st.secrets["ANTHROPIC_API_KEY"]
    except (KeyError, FileNotFoundError):
        return None


# ─── Estado da sessão ────────────────────────────────────────────────────────

def init_session():
    """Inicializa o estado da sessão na primeira execução."""
    if "workbook_bytes" not in st.session_state:
        # Criar planilha template nova
        st.session_state.workbook_bytes = create_template().read()

    if "display_messages" not in st.session_state:
        st.session_state.display_messages = []  # [{role, content}]

    if "api_messages" not in st.session_state:
        st.session_state.api_messages = []  # Histórico completo para a API

    if "project_name" not in st.session_state:
        st.session_state.project_name = ""

    if "last_saved" not in st.session_state:
        st.session_state.last_saved = None

    if "draft_count" not in st.session_state:
        st.session_state.draft_count = 0


def get_workbook() -> openpyxl.Workbook:
    """Carrega o workbook atual do session_state."""
    return load_workbook_from_bytes(st.session_state.workbook_bytes)


def save_workbook(wb: openpyxl.Workbook):
    """Salva o workbook atualizado no session_state."""
    st.session_state.workbook_bytes = save_workbook_to_bytes(wb)
    st.session_state.last_saved = datetime.datetime.now()


# ─── Sidebar ─────────────────────────────────────────────────────────────────

def render_sidebar():
    with st.sidebar:
        # Logo / marca
        logo_uri = _logo_data_uri()
        logo_img = f'<img src="{logo_uri}" alt="Gabriela Lendecker" class="brand-logo" />' if logo_uri else ""
        st.markdown(
            f"""
            <div class="brand-header">
              {logo_img}
              <span class="brand-name">Gabriela Lendecker</span>
            </div>
            <div class="brand-product">Arqui Specs</div>
            <div class="brand-sub">Ferramenta Interna</div>
            """,
            unsafe_allow_html=True,
        )
        st.divider()

        # Nome do projeto
        project_name = st.text_input(
            "Nome do projeto",
            value=st.session_state.project_name,
            placeholder="",
            key="project_name_input",
        )
        if project_name != st.session_state.project_name:
            st.session_state.project_name = project_name
            wb = get_workbook()
            update_project_info(wb, nome=project_name)
            save_workbook(wb)

        st.divider()

        # ── Upload de planilha existente ────────────────────────────────
        st.markdown("**Planilha**")
        if st.session_state.get("last_uploaded_id"):
            # Já há uma planilha carregada: mostrar opção de trocar
            st.caption("Planilha carregada")
            if st.button("Trocar planilha", use_container_width=True, help="Remove a planilha atual e permite carregar outra", type="primary"):
                st.session_state.last_uploaded_id = None
                st.rerun()
        else:
            uploaded = st.file_uploader(
                "Carregar .xlsx existente",
                type=["xlsx"],
                label_visibility="collapsed",
                key="uploader",
            )
            if uploaded:
                try:
                    wb_test = load_workbook_from_bytes(uploaded.getvalue())
                    st.session_state.workbook_bytes = uploaded.getvalue()
                    st.session_state.last_uploaded_id = uploaded.file_id
                    info = get_project_info(wb_test)
                    st.session_state.project_name = info["nome"]
                    st.session_state.display_messages = []
                    st.session_state.api_messages = []
                    st.success("Planilha carregada!")
                    st.rerun()
                except Exception as e:
                    st.error(f"Arquivo inválido: {e}")

        # ── Download ──────────────────────────────────────────────────────
        wb = get_workbook()
        info = get_project_info(wb)
        filename = f"Memorial_{info['nome'].replace(' ', '_').replace('/', '-')}.xlsx"

        st.download_button(
            label="Baixar .xlsx",
            data=st.session_state.workbook_bytes,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        # Auto-save info (SCRUM-62)
        if st.session_state.last_saved:
            ts = st.session_state.last_saved.strftime("%H:%M:%S")
            st.caption(f"Último save: {ts}")

        st.divider()

        # ── Nova planilha / limpar conversa ───────────────────────────────────
        if st.button("Nova planilha", use_container_width=True, help="Começa do zero", type="primary"):
            st.session_state.workbook_bytes = create_template(
                project_name=st.session_state.project_name
            ).read()
            st.session_state.display_messages = []
            st.session_state.api_messages = []
            st.session_state.last_saved = None
            st.rerun()
        if st.button("Limpar chat", use_container_width=True, help="Mantém planilha", type="primary"):
            st.session_state.display_messages = []
            st.session_state.api_messages = []
            st.rerun()

        st.divider()

        # ── Info da planilha atual ─────────────────────────────────────────
        wb = get_workbook()
        data = get_sheet_data(wb)
        n_items = len(data)
        n_ambientes = len({r["ambiente"] for r in data if r["ambiente"]})
        st.markdown(f"**{n_items} itens · {n_ambientes} ambientes/abas**")

        # Indicador da chave da API
        api_key = get_api_key()
        if not api_key:
            st.error("API key não configurada.\nCrie `.streamlit/secrets.toml`.")

        st.divider()
        st.caption(f"Modelo: {MODEL}")
        st.caption("v1.0 · Arqui Specs")


# ─── Processamento de mensagem ───────────────────────────────────────────────

def process_message(user_input: str, api_key: str, preview_placeholder):
    """
    Processa uma mensagem do usuário:
    1. Mostra mensagem do usuário
    2. Chama Claude com streaming + tool use
    3. Se houve tool use, salva planilha atualizada
    4. Atualiza preview ao vivo
    5. Salva mensagem no histórico
    """
    wb = get_workbook()

    # Adicionar mensagem ao histórico de display
    st.session_state.display_messages.append({"role": "user", "content": user_input})

    # Construir mensagem para a API (com contexto da planilha)
    api_user_msg = build_user_message(user_input, wb)
    st.session_state.api_messages.append(api_user_msg)

    # Streaming da resposta
    with st.chat_message("assistant"):
        placeholder = st.empty()
        full_response = ""
        workbook_updated = False
        error_msg = None

        for chunk, updated_wb, err in stream_response(
            api_key=api_key,
            api_messages=st.session_state.api_messages,
            workbook=wb,
        ):
            if err:
                error_msg = err
                break
            if updated_wb:
                # Workbook foi modificado pelo tool use
                save_workbook(updated_wb)
                workbook_updated = True
            if chunk:
                # Limpa o marcador "*Atualizando planilha...*" da exibição final
                full_response += chunk
                # Remover o marcador de loading do texto final exibido
                display_text = full_response.replace("*Atualizando planilha...*\n\n", "")
                placeholder.markdown(display_text + "▌")

        if error_msg:
            placeholder.error(error_msg)
            full_response = error_msg
        else:
            display_text = full_response.replace("*Atualizando planilha...*\n\n", "")
            placeholder.markdown(display_text)
            full_response = display_text

    # Salvar resposta no histórico
    st.session_state.display_messages.append({"role": "assistant", "content": full_response})
    st.session_state.api_messages.append({"role": "assistant", "content": full_response})

    # Atualizar preview se planilha mudou (SCRUM-54)
    if workbook_updated:
        render_preview(preview_placeholder, get_workbook())
        # Auto-increment draft counter (SCRUM-62)
        st.session_state.draft_count += 1


# ─── Preview da planilha (SCRUM-54) ──────────────────────────────────────────

def render_preview(container, wb: openpyxl.Workbook):
    """Renderiza o preview da planilha no container fornecido."""
    with container:
        df = get_preview_dataframe(wb)
        info = get_project_info(wb)

        if df.empty:
            st.info("Comece descrevendo os itens do projeto no chat ao lado.")
            st.markdown("**Exemplo:** *cozinha: cuba Tramontina inox 50x40, torneira Docol gourmet preta*")
            return

        # Cabeçalho do projeto
        st.markdown(f"**{info['nome']}**")
        if info["cliente"] and info["cliente"] != "(definir)":
            st.caption(f"Cliente: {info['cliente']} · {info['emissao']} · {info['versao']}")

        # Tabela — usa tema nativo do Streamlit (config.toml define as cores)
        st.dataframe(
            df,
            use_container_width=True,
            height=min(600, 60 + len(df) * 38),
            hide_index=True,
        )

        # Rodapé de contagem
        n = len(df)
        n_amb = df["AMBIENTE"].nunique() if "AMBIENTE" in df.columns else 0
        st.caption(f"{n} itens · {n_amb} ambientes/abas")


# ─── App principal ───────────────────────────────────────────────────────────

def main():
    # 1. Auth
    if not check_auth():
        return

    # 2. Init
    init_session()
    api_key = get_api_key()

    # 3. Sidebar
    render_sidebar()

    # 4. Layout: 2 colunas
    col_chat, col_preview = st.columns([1, 1], gap="large")

    # 5. Coluna de chat (esquerda)
    with col_chat:
        st.markdown("### Conversa")

        # Container com altura fixa e scrollbar — evita que a página cresça indefinidamente
        chat_container = st.container(height=580)

        with chat_container:
            # Mensagem de boas-vindas se chat vazio
            if not st.session_state.display_messages:
                with st.chat_message("assistant"):
                    st.markdown(
                        "Olá! Pronta para montar o memorial.\n\n"
                        "Descreva os itens em linguagem natural — ambiente por ambiente:\n\n"
                        "> *cozinha: cuba Tramontina inox 50x40, torneira Docol gourmet preta, "
                        "cooktop Electrolux IE80P, 2 arandelas para suíte casal*"
                    )

            # Histórico de mensagens
            for msg in st.session_state.display_messages:
                with st.chat_message(msg["role"]):
                    st.markdown(msg["content"])

        # Input fora do container — fica fixo abaixo da área de scroll
        user_input = st.chat_input(
            placeholder="Descreva os itens do projeto...",
            disabled=(api_key is None),
        )

    # 6. Coluna de preview (direita) — sempre renderiza com estado atual
    with col_preview:
        st.markdown("### Planilha ao vivo")
        preview_placeholder = st.container()
        wb_current = get_workbook()
        render_preview(preview_placeholder, wb_current)

    # 7. Processar input (após ambas as colunas serem definidas)
    if user_input and api_key:
        with col_chat:
            with st.chat_message("user"):
                st.markdown(user_input)
        process_message(user_input, api_key, preview_placeholder)
        # Re-renderiza o app inteiro para que a sidebar (download button, contadores)
        # reflita o workbook atualizado. Sem isso, o botão de download captura
        # os bytes da renderização anterior (planilha vazia).
        st.rerun()

    elif user_input and not api_key:
        st.error(
            "Configure a API key em `.streamlit/secrets.toml` antes de usar.\n"
            "Copie `.streamlit/secrets.toml.example` e preencha com sua chave."
        )


if __name__ == "__main__":
    main()
